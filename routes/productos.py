from flask import Blueprint, request, jsonify
from flask_login import login_required, current_user
from database import db
from models import Producto
import io

productos_bp = Blueprint('productos', __name__)

@productos_bp.route('/', methods=['GET'])
@login_required
def listar():
    productos = Producto.query.order_by(Producto.categoria, Producto.nombre).all()
    return jsonify([p.to_dict() for p in productos])

@productos_bp.route('/', methods=['POST'])
@login_required
def crear():
    if current_user.rol != 'admin':
        return jsonify({'error': 'Solo el admin puede agregar productos'}), 403
    data = request.json
    if not data.get('codigo') or not data.get('nombre'):
        return jsonify({'error': 'Código y nombre son requeridos'}), 400
    if Producto.query.filter_by(codigo=data['codigo']).first():
        return jsonify({'error': 'El código ya existe'}), 400
    p = Producto(
        codigo=data['codigo'], nombre=data['nombre'],
        categoria=data.get('categoria', 'General'),
        stock=int(data.get('stock', 0)), minimo=int(data.get('minimo', 0))
    )
    db.session.add(p)
    db.session.commit()
    return jsonify(p.to_dict()), 201

@productos_bp.route('/importar', methods=['POST'])
@login_required
def importar():
    if current_user.rol != 'admin':
        return jsonify({'error': 'Solo el admin puede importar productos'}), 403

    if 'archivo' not in request.files:
        return jsonify({'error': 'No se envió ningún archivo'}), 400

    archivo = request.files['archivo']
    if not archivo.filename.endswith('.xlsx'):
        return jsonify({'error': 'El archivo debe ser .xlsx'}), 400

    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(archivo.read()))
        ws = wb.active

        creados = 0
        omitidos = 0
        errores = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            codigo, nombre, categoria, stock, minimo = (list(row) + [None]*5)[:5]
            if not codigo or not nombre:
                errores.append(f'Fila sin código o nombre: {row}')
                continue
            codigo = str(codigo).strip()
            nombre = str(nombre).strip()
            categoria = str(categoria).strip() if categoria else 'General'
            stock = int(stock) if stock else 0
            minimo = int(minimo) if minimo else 0

            if Producto.query.filter_by(codigo=codigo).first():
                omitidos += 1
                continue

            db.session.add(Producto(codigo=codigo, nombre=nombre, categoria=categoria, stock=stock, minimo=minimo))
            creados += 1

        db.session.commit()
        return jsonify({'creados': creados, 'omitidos': omitidos, 'errores': errores})

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@productos_bp.route('/plantilla', methods=['GET'])
@login_required
def plantilla():
    if current_user.rol != 'admin':
        return jsonify({'error': 'Acceso denegado'}), 403
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from flask import send_file

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Productos'
    headers = ['codigo', 'nombre', 'categoria', 'stock', 'minimo']
    header_fill = PatternFill('solid', start_color='00E5A0')
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True, color='000000')
        cell.alignment = Alignment(horizontal='center')
    ejemplos = [('PRD-001', 'Ejemplo Producto 1', 'Categoría A', 100, 10),
                ('PRD-002', 'Ejemplo Producto 2', 'Categoría B', 50, 5)]
    for row, data in enumerate(ejemplos, 2):
        for col, val in enumerate(data, 1):
            ws.cell(row=row, column=col, value=val)
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='plantilla_productos.xlsx')

@productos_bp.route('/<int:id>', methods=['PUT'])
@login_required
def actualizar(id):
    if current_user.rol != 'admin':
        return jsonify({'error': 'Solo el admin puede editar productos'}), 403
    p = Producto.query.get_or_404(id)
    data = request.json
    p.nombre    = data.get('nombre', p.nombre)
    p.categoria = data.get('categoria', p.categoria)
    p.minimo    = int(data.get('minimo', p.minimo))
    db.session.commit()
    return jsonify(p.to_dict())

@productos_bp.route('/<int:id>', methods=['DELETE'])
@login_required
def eliminar(id):
    if current_user.rol != 'admin':
        return jsonify({'error': 'Solo el admin puede eliminar productos'}), 403
    p = Producto.query.get_or_404(id)
    db.session.delete(p)
    db.session.commit()
    return jsonify({'ok': True})
